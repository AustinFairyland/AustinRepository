# coding: utf8
""" 
@software: PyCharm
@author: Lionel Johnson
@contact: https://fairy.host
@organization: https://github.com/FairylandFuture
@since: 2024-04-23 16:01:43 UTC+8
"""

import yaml
import os
import openpyxl

from typing import TypeVar

_TypeFileBaseService = TypeVar("_TypeFileBaseService", bound="FileBaseService")
_TypeYamlFileService = TypeVar("_TypeYamlFileService", bound="YamlFileService")
_TypeExcelFileService = TypeVar("_TypeExcelFileService", bound="ExcelFileService")


class FileBaseService(object):
    """ 文件服务基础类 """

    def __init__(self, path):
        self.__path = path

        self.validate_path(path)

    @property
    def path(self):
        return self.__path

    def validate_path(self, value):
        if not isinstance(value, basestring):
            raise TypeError("参数必须是字符串")


class YamlFileService(FileBaseService):
    """ YAML 文件服务 """

    def __init__(self, yaml_path):
        super(YamlFileService, self).__init__(yaml_path)

        self.validate_yaml_path(yaml_path)

    def validate_yaml_path(self, value):  # type: (str) -> None
        if not value.endswith('.yaml'):
            raise TypeError("YAML文件类型错误")

    def read(self):  # type: () -> dict
        with open(self.path, 'r') as binary_io:
            content = yaml.safe_load(binary_io)
        return content


class ExcelFileService(FileBaseService):
    """ Excel 文件服务 """

    def __init__(self, excel_path):
        super(ExcelFileService, self).__init__(excel_path)

        self.validate_excel_path(excel_path)

        self.__workbook = openpyxl.load_workbook(self.path)
        self.__worksheet = self.__workbook.worksheets.__getitem__(0)

    def validate_excel_path(self, value):  # type: (str) -> None
        if not value.endswith('.xlsx'):
            raise TypeError("Excel文件类型错误")

    @property
    def workbook(self):
        return self.__workbook

    @property
    def worksheet(self):
        return self.__worksheet

    @worksheet.setter
    def worksheet(self, value):
        if not isinstance(value, int):
            raise TypeError("参数类型必须是整型")

        value -= 1

        try:
            self.__worksheet = self.__workbook.worksheets.__getitem__(value)
        except IndexError:
            raise IndexError("工作表超出最大索引(没有这个Sheet页)")

    def get_all_data(self):
        data = list()
        for row in self.__worksheet.iter_rows(values_only=True):
            data.append(row)

        return tuple(data)

    def iter_all_data_values(self):
        for row in self.__worksheet.iter_rows(values_only=True):
            yield row
