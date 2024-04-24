# coding: utf8
""" 
@software: PyCharm
@author: Lionel Johnson
@contact: https://fairy.host
@organization: https://github.com/FairylandFuture
@since: 2024-04-22 18:57:40 UTC+8
"""

import os
import openpyxl

from typing import TypeVar

_TypeExcelBaseService = TypeVar("_TypeExcelBaseService", bound="ExcelBaseService")


class ExcelBaseService(object):

    def __init__(self, path):
        self.__path = path
        self.__load_workbook = openpyxl.load_workbook(self.__path)
        self.__load_sheetbook = self.__load_workbook.worksheets.__getitem__(0)

    @property
    def workbook(self):
        return self.__load_workbook

    @property
    def sheetbook(self):
        return self.__load_sheetbook

    @sheetbook.setter
    def sheetbook(self, value):
        if not isinstance(value, int):
            raise TypeError("参数类型必须是整形")

        value -= 1

        try:
            self.__load_sheetbook = self.__load_workbook.worksheets.__getitem__(value)
        except IndexError:
            raise IndexError("工作表超出最大索引")

    def get_all_data(self):
        data = list()
        for row in self.__load_sheetbook.iter_rows(values_only=True):
            data.append(row)

        return tuple(data)

    def iter_all_data_values(self):
        for row in self.__load_sheetbook.iter_rows(values_only=True):
            yield row
